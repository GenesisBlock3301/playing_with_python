import openpyxl


def common_return(file, sheet_name):
    work_book = openpyxl.load_workbook(file)
    return work_book.get_sheet_by_name(sheet_name), work_book


def get_row_count(file, sheet_name):
    sheet, _ = common_return(file, sheet_name)
    return sheet.max_row


def get_column_count(file, sheet_name):
    sheet, _ = common_return(file, sheet_name)
    return sheet.max_column


def read_data(file, sheet_name, rownum, columno):
    sheet, _ = common_return(file, sheet_name)
    return sheet.cell(row=rownum, column=columno)


def write_data(file, sheet_name, rownum, columnno, data):
    sheet, work_book = common_return(file, sheet_name)
    sheet.cell(row=rownum, column=columnno).value = data
    work_book.save(file)


if __name__ == "__main__":
    FILE = "./sample1.xlsx"
    print(get_row_count(file=FILE, sheet_name="Sheet1"))
    print(get_column_count(file=FILE, sheet_name="Sheet1"))
    print(read_data(file=FILE, sheet_name="Sheet1", rownum=1, columno=2).value)
