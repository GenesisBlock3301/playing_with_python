from openpyxl import load_workbook

workbook = load_workbook(filename="./hello_world.xlsx")
sheets = workbook.sheetnames
sheet = workbook[sheets[0]]

rows = len(sheet["A"])
column = len(sheet[1])

for i in range(1, rows + 1):
   for j in range(1, column + 1):
      if i == 2 and j == 2:
         sheet.cell(row=i,column=j).value = "OK"
      print(sheet.cell(row=i, column=j).value, end="\n")
   print("==================")
